import streamlit as st
import sys
import datetime
import numpy as np
from PIL import Image
import pandas as pd
from st_aggrid import AgGrid, GridOptionsBuilder
from st_aggrid.shared import GridUpdateMode
from babel.dates import format_date, format_datetime, format_time
import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl.styles import Font, Fill
from openpyxl.styles.borders import Border, Side
from io import BytesIO
import os

@st.cache_data()
def load_data(file,option1,option2,option3,d):
    
    book=load_workbook('R.xlsx')
    book1=load_workbook(file)
    sheet=book.active
    sheet1=book1.active
    sheet1.delete_rows(1,3)
    sheet1.delete_cols(6,4)
    mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleLight1',
                                                      showRowStripes=True)
    table = openpyxl.worksheet.table.Table(ref='A1:E'+str(len(sheet1['E'])),
										displayName='xf',
										tableStyleInfo=mediumStyle)
    sheet1.add_table(table)
    
    #sheet1['E1'].value=option1
    sheet1.column_dimensions['A'].width = 18
    sheet1.column_dimensions['B'].width = 40
    sheet1.column_dimensions['C'].width = 27
    sheet1.column_dimensions['D'].width = 20
    sheet1.column_dimensions['E'].width = 18
    font = Font(size=12)
    sheet1['C1'].value=str(option3)
    for i, row in enumerate(sheet1):
        sheet1['A'+str(i+1)].font=Font(size=12)
        sheet1.row_dimensions[i].height = 15
        sheet1['B'+str(i+1)].font=Font(size=12)
        sheet1.row_dimensions[i].height = 15
        sheet1['C'+str(i+1)].font=Font(size=12)
        sheet1.row_dimensions[i].height = 15
        sheet1['D'+str(i+1)].font=Font(size=12)
        sheet1.row_dimensions[i].height = 15
        sheet1['E'+str(i+1)].font=Font(size=12)
        sheet1.row_dimensions[i].height = 15
    a=len(sheet1['E'])
    b=len(sheet['A'])
    thin_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    sheet1['D1'].value=str(d)
    for i, row in enumerate(sheet):
        
        sheet1['A'+str(i+a+1)].value=sheet['A'+str(i+1)].value
        sheet1['A'+str(i+a+1)].font=Font(size=14)
        sheet1['A'+str(i+a+1)].border = thin_border
        sheet1['B'+str(i+a+1)].value=sheet['B'+str(i+1)].value
        sheet1['B'+str(i+a+1)].font=Font(size=14)
        sheet1['B'+str(i+a+1)].border = thin_border
        sheet1['C'+str(i+a+1)].value=sheet['C'+str(i+1)].value
        sheet1['C'+str(i+a+1)].font=Font(size=14)
        sheet1['C'+str(i+a+1)].border = thin_border
        sheet1['D'+str(i+a+1)].value=sheet['D'+str(i+1)].value
        sheet1['D'+str(i+a+1)].font=Font(size=14)
        sheet1['D'+str(i+a+1)].border = thin_border
        sheet1['E'+str(i+a+1)].value=sheet['E'+str(i+1)].value
        sheet1['E'+str(i+a+1)].font=Font(size=14)
        sheet1['E'+str(i+a+1)].border = thin_border
        sheet1.row_dimensions[i+a].height = 22
    sheet1.merge_cells('A'+str(a+1)+':B'+str(a+1))
    sheet1.merge_cells('A'+str(a+2)+':B'+str(a+2))
    sheet1.merge_cells('A'+str(a+3)+':B'+str(a+3))
    sheet1.merge_cells('A'+str(a+4)+':B'+str(a+4))
    sheet1.merge_cells('A'+str(a+5)+':B'+str(a+5))
    sheet1.merge_cells('A'+str(a+6)+':B'+str(a+6))
    sheet1.merge_cells('D'+str(a+1)+':E'+str(a+1))
    sheet1.merge_cells('D'+str(a+2)+':E'+str(a+6))
    sheet1.merge_cells('A'+str(a+7)+':B'+str(a+7))
    sheet1.merge_cells('A'+str(a+8)+':B'+str(a+8))
    sheet1.merge_cells('C'+str(a+7)+':E'+str(a+7))
    sheet1.merge_cells('C'+str(a+8)+':E'+str(a+8))
    sheet1.merge_cells('C'+str(a+9)+':E'+str(a+9))
    x = datetime.datetime.now()
    sheet1['C'+str(a+9)]=x.strftime("%H:%M:%S")
    sheet1['B1'].value=str(option2)
    
    book1.save(option1+'.xlsx')
    #book1.save(r""+t+"\\"+option1+'.xlsx')

    #df =  pd.read_excel (option1+'.xlsx')
    #df.to_excel(t+'\\'+option1+'.xlsx', index=False)
   
    
    
    
hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)


st.title('BACK OFFICE SARL ANDROMED DISTRIBUTION -ALGER CENTRE- 2023')
st.text('Application Créer Par: ALLOUCHE KENZA')

st.session_state["Page1"]="Went Page1"

st.divider()
st.header('DATA BASE:')
uploaded_files = st.file_uploader("IMPORTER LE BON DE CHARGEMENT ", accept_multiple_files=True)
for uploaded_file in uploaded_files:
    bytes_data = uploaded_file.read()
    print (type(bytes_data))
    st.write("NOM DE FICHIER:", uploaded_file.name)
st.divider()
col1, col2,col3 = st.columns(3)


with col1:
    option1 = st.selectbox(
        "ROUTE :",
        ("PS16F01",
         "PS16F02",
         "PS16F03",
         "PS16F04",
         "PS16F05",
         "PS16F06",
         "PS16F07",
         "PS16F08",
         "PS16F09",
         "PS16F10",
         "PS16F11",
         "PS16F12",
         "PS16F13",
         "PS16F14",
         "PS16F15",
         "PS16F16",
         "PS16F17",),
        key="v1",
    )

with col2:
    option2 = st.selectbox(
        "VENDEUR",
        ("MANSOUR HICHEM",
        "TOUADI MORAD",
        'GUERRASSI HOUSSEM',
        "KRELIFAOUI YOUCEF",
        "LOULANSSA KHALD",
        "KADEM ISLAM",
        "GUERRIDA M HAMED",
        'ABDESSELAM FARID',
        'OUBOUCHOU KAMEL',
        'BRAHIMI BOUBEKEUR',
        'YAHIAOUI YOUCEF',
        'BENIGHINE MEROUANE',
        'BOUDALI IMAD EDDINE',
        'EL KADI ABDELMADJID MADJED',
        'AMMAM ABDELKRIM',
        'OUARTI YACINE',
        'DIAB ISMAIL',
        'HAMADACHE SOFIANE',
        'AHMED MENSOUR'
        ),
        key="v2",
        
    )
with col3:
    option3 = st.selectbox(
        "LIVREUR",
        (
           'ZOUBRI AMINE',
            'BELOUDINA RACHID',
            'OUAHIB ABDERRAHMANE',
            'LAOUANA FOUAD',
            'SAID HADJAZ',
            'LAOUAR ZAKARIA',
            'ALIOUA AYOUB',
            'MOHAMEDI MOKHTAR',
            'BEN TEFRAOUINE FAHIM',
            'ABDELLOUCHE NAZIM',
            'LEKBEDJ ABBES',
            'ARABI ABDELLAH',
            'EL KADI ABDELMADJID MADJED',
            'TOUHAMI MOHAMED',
            'REZZOUG IMAD',
            'ESSEGHIR HOCINE',
            'LEGAB BILEL',
            'ACILA ABDELLAH',
            'BENNOUI HACHEM',
            'BENBOUZID AYMEN '),
        key="v3",
        
    )
st.divider()

d = st.date_input("CHOISIR UNE DATE LIVRAISON",datetime.datetime.now())
format_date(d, locale='en')
st.write('DATE LIVRAISON:', d)
st.divider()

if st.button('EXECUTE'):
    

# Join various path components
    
    #fill=
    load_data(uploaded_file,option1,option2,option3,d)
    
    with open(option1+'.xlsx', "rb") as template_file:
        template_byte = template_file.read()

    st.download_button(label="Download BON ",
                        data=template_byte,
                        file_name=option1+'.xlsx',
                        )

